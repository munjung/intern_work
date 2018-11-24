import openpyxl

class ExcelFormat():

    def __init__(self, filename):
        self.filename = filename

    def __enter__(self):
        self.wb = openpyxl.load_workbook(self.filename)
        self.ws = self.wb.active
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.wb.save(self.filename)
        self.wb.close()

    def set_value(self, row, col, value):
        self.ws.cell(row=row, column=col).value = value

    def get_value(self, row, col):
        return self.ws.cell(row=row, column=col)
